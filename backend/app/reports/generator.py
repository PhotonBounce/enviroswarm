"""Report generation utilities for ENViroSwarm."""

import csv
import os
import tempfile
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional
from uuid import UUID

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import SensorReading, SensorStation


class ReportGenerator:
    """Generate PDF, CSV, and Excel reports from sensor data."""

    def __init__(self, db: AsyncSession, user_id: UUID):
        self.db = db
        self.user_id = user_id

    async def _fetch_readings(
        self,
        station_id: Optional[UUID],
        start: Optional[datetime],
        end: Optional[datetime],
    ) -> List[Dict[str, Any]]:
        """Fetch readings for the report, filtered by station and date range."""
        stmt = (
            select(SensorReading, SensorStation.name.label("station_name"))
            .join(SensorStation)
            .where(SensorStation.user_id == self.user_id)
            .where(SensorStation.deleted_at.is_(None))
            .where(SensorReading.deleted_at.is_(None))
            .order_by(SensorReading.timestamp.asc())
        )
        if station_id is not None:
            stmt = stmt.where(SensorReading.station_id == station_id)
        if start is not None:
            stmt = stmt.where(SensorReading.timestamp >= start)
        if end is not None:
            stmt = stmt.where(SensorReading.timestamp <= end)

        result = await self.db.execute(stmt)
        rows = []
        for row in result.all():
            reading = row[0]
            rows.append({
                "station_name": row.station_name or "Unknown",
                "sensor_type": reading.sensor_type,
                "value": float(reading.value),
                "unit": reading.unit,
                "timestamp": reading.timestamp.isoformat(),
            })
        return rows

    async def generate_csv(
        self,
        station_id: Optional[UUID],
        start: Optional[datetime],
        end: Optional[datetime],
        output_path: str,
    ) -> str:
        """Generate a CSV report and return the file path."""
        rows = await self._fetch_readings(station_id, start, end)
        with open(output_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(
                f,
                fieldnames=["station_name", "sensor_type", "value", "unit", "timestamp"],
            )
            writer.writeheader()
            writer.writerows(rows)
        return output_path

    async def generate_excel(
        self,
        station_id: Optional[UUID],
        start: Optional[datetime],
        end: Optional[datetime],
        output_path: str,
    ) -> str:
        """Generate an Excel (.xlsx) report and return the file path."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
        except ImportError as exc:
            raise RuntimeError("openpyxl is required for Excel generation") from exc

        rows = await self._fetch_readings(station_id, start, end)
        wb = Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet("Report")
        ws.title = "Sensor Data"

        headers = ["Station", "Sensor Type", "Value", "Unit", "Timestamp"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        for row in rows:
            ws.append([
                row["station_name"],
                row["sensor_type"],
                row["value"],
                row["unit"],
                row["timestamp"],
            ])

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width

        wb.save(output_path)
        return output_path

    async def generate_pdf(
        self,
        station_id: Optional[UUID],
        start: Optional[datetime],
        end: Optional[datetime],
        output_path: str,
    ) -> str:
        """Generate a PDF report and return the file path."""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
        except ImportError as exc:
            raise RuntimeError("reportlab is required for PDF generation") from exc

        rows = await self._fetch_readings(station_id, start, end)
        doc = SimpleDocTemplate(output_path, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()

        title = "ENViroSwarm Sensor Report"
        elements.append(Paragraph(title, styles["Title"]))
        elements.append(Spacer(1, 12))

        meta_lines = [
            f"Generated: {datetime.now(timezone.utc).isoformat()}",
        ]
        if start and end:
            meta_lines.append(f"Period: {start.isoformat()} to {end.isoformat()}")
        if station_id:
            meta_lines.append(f"Station ID: {station_id}")
        for line in meta_lines:
            elements.append(Paragraph(line, styles["Normal"]))
        elements.append(Spacer(1, 12))

        table_data = [["Station", "Sensor", "Value", "Unit", "Timestamp"]]
        for row in rows[:1000]:  # Limit PDF rows
            table_data.append([
                row["station_name"],
                row["sensor_type"],
                str(row["value"]),
                row["unit"],
                row["timestamp"],
            ])

        if len(rows) > 1000:
            elements.append(Paragraph(
                f"Note: Showing first 1000 of {len(rows)} records.",
                styles["Normal"],
            ))
            elements.append(Spacer(1, 6))

        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563eb")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#f8fafc")),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(table)
        doc.build(elements)
        return output_path

    async def generate(
        self,
        report_format: str,
        station_id: Optional[UUID],
        start: Optional[datetime],
        end: Optional[datetime],
        output_dir: str,
    ) -> str:
        """Dispatch to the correct generator based on format."""
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        ext_map = {"pdf": "pdf", "csv": "csv", "excel": "xlsx"}
        ext = ext_map.get(report_format, report_format)
        filename = f"report_{timestamp}.{ext}"
        output_path = os.path.join(output_dir, filename)
        os.makedirs(output_dir, exist_ok=True)

        if report_format == "csv":
            return await self.generate_csv(station_id, start, end, output_path)
        if report_format == "excel":
            return await self.generate_excel(station_id, start, end, output_path)
        if report_format == "pdf":
            return await self.generate_pdf(station_id, start, end, output_path)
        raise ValueError(f"Unsupported report format: {report_format}")
